# -*- coding: utf-8 -*-
"""
신뢰성분석 (가속시험 통합계산기) - Streamlit 웹앱
원본 데스크톱 프로그램(신뢰성분석.py, tkinter)의 화면 구조를 최대한 그대로 재현했습니다.
- 좌측: 입력값 / 우측: 결과(파란 박스) + 계산공식
- 탭① 온도가속(Arrhenius)
- 탭② 온습도가속(Arrhenius-Peck)
- 탭③ 열피로가속(Thermal Cycling)
- 탭④ Weibull 시험시간비 계산기 + 참고 DB
- 탭⑤ 수명데이터 분석 (Weibull MLE) + 확률도표
"""
import io
import math
import pandas as pd
import numpy as np
import streamlit as st
import matplotlib.pyplot as plt

import reliability_calc as R

st.set_page_config(page_title="신뢰성분석 - 가속시험 통합계산기", layout="wide")

# ------------------------------------------------------------------
# 원본 tkinter 스타일(연한 파란 결과박스 / 공식박스 / 빨간 경고문) 재현용 CSS
# ------------------------------------------------------------------
st.markdown("""
<style>
.result-box{
    background-color:#eef2fb; color:#1a3fa0; padding:14px 16px; border-radius:6px;
    font-weight:700; font-size:15px; white-space:pre-wrap; line-height:1.7;
    border:1px solid #d7e0f5;
}
.formula-box{
    background-color:#f5f5f5; padding:12px 14px; border-radius:6px;
    white-space:pre-wrap; font-size:13px; line-height:1.6; color:#333;
    border:1px solid #e2e2e2;
}
.detail-box{
    font-size:13px; color:#333; white-space:pre-wrap; line-height:1.6; padding:4px 2px;
}
.note-red{color:#a33; font-size:13px;}
.note-gray{color:#666; font-size:12px;}
div.stButton > button[kind="primary"]{
    background-color:#2f6fed; color:white; font-weight:700;
}
div.stButton > button[kind="primary"]:hover{
    background-color:#255bc4; color:white;
}
</style>
""", unsafe_allow_html=True)

st.markdown("## 가속시험 통합계산기")

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    " ① 온도가속(Arrhenius) ",
    " ② 온습도가속(Arrhenius-Peck) ",
    " ③ 열피로가속(Thermal Cycling) ",
    " ④ Weibull 시험시간비 ",
    " ⑤ 수명데이터 분석 ",
])


def _fmt_num(v, nd=2):
    try:
        return f"{v:,.{nd}f}"
    except Exception:
        return str(v)


def excel_download_button(label, key, sheet_builder, filename):
    """엑셀로 내보내기 버튼: 클릭 시 세션에 저장된 결과로 xlsx 생성 후 다운로드"""
    if st.button(label, key=key + "_btn"):
        if st.session_state.get(key + "_ready"):
            wb = sheet_builder()
            buf = io.BytesIO()
            wb.save(buf)
            st.session_state[key + "_bytes"] = buf.getvalue()
        else:
            st.warning("먼저 [계산 실행]을 눌러주세요.")
    if st.session_state.get(key + "_bytes"):
        st.download_button("⬇ 다운로드: " + filename, data=st.session_state[key + "_bytes"],
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=key + "_dl")


# ====================================================================
# 탭① 온도가속 (Arrhenius Model)
# ====================================================================
with tab1:
    left, right = st.columns([2, 3])
    with left:
        st.markdown("#### ① 온도가속 (Arrhenius Model)")
        st.markdown("<span class='note-red'>※ 사이클(열충격) 시험이 아닌, 단일 온도 유지 시험(정특성)에 적용합니다.</span>",
                    unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**입력값**")
            c1, c2 = st.columns([3, 1])
            with c1:
                E = st.number_input("활성화에너지 E (eV)", value=0.8, step=0.01, format="%.3f", key="t1_E")
            with c2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                with st.popover("DB에서 찾기"):
                    q = st.text_input("부품명 검색", key="t1_dbq")
                    rows = [r for r in R.EA_DB if q.lower() in r[0].lower()] if q else R.EA_DB
                    st.dataframe(pd.DataFrame(rows, columns=["부품", "활성화에너지(eV)", "출처", "비고"]),
                                 height=280, use_container_width=True)
                    names = [r[0] for r in rows]
                    sel = st.selectbox("적용할 항목", ["-"] + names, key="t1_dbsel")
                    if st.button("선택값 적용", key="t1_dbapply"):
                        if sel != "-":
                            row = next(r for r in rows if r[0] == sel)
                            ev_text = str(row[1]).split("~")[0].split("-")[0].replace("eV", "").strip()
                            try:
                                st.session_state["t1_E"] = float(ev_text)
                                st.success(f"E = {ev_text} 적용됨 (다시 열어 확인)")
                                st.rerun()
                            except ValueError:
                                st.warning("이 항목은 단일 숫자값이 아니라 자동입력이 어렵습니다. 직접 입력해주세요.")
            Tref = st.number_input("대표(목표) 온도 Tref (℃)", value=125.0, step=1.0, key="t1_Tref")

        with st.container(border=True):
            st.markdown("**시험 조건 (온도 / 시간) 추가**")
            cc1, cc2, cc3 = st.columns([1, 1, 1])
            ta = cc1.number_input("조건온도(℃)", value=100.0, step=1.0, key="t1_ta")
            tt = cc2.number_input("조건시간(h)", value=200.0, step=1.0, key="t1_tt")
            with cc3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("추가 +", key="t1_add"):
                    st.session_state.setdefault("t1_conditions", []).append((ta, tt))

        with st.container(border=True):
            st.markdown("**조건 목록**")
            conds = st.session_state.get("t1_conditions", [])
            del_idx = []
            if conds:
                df = pd.DataFrame(conds, columns=["조건온도(℃)", "조건시간(h)"])
                st.dataframe(df, use_container_width=True, height=180)
                del_idx = st.multiselect("삭제할 행 선택", list(range(len(conds))), key="t1_delsel",
                                          format_func=lambda i: f"{i}: {conds[i][0]}℃ / {conds[i][1]}h")
            else:
                st.caption("추가된 조건이 없습니다.")
            bc1, bc2 = st.columns(2)
            if bc1.button("선택 삭제 🗑", key="t1_delbtn", use_container_width=True):
                if conds and del_idx:
                    st.session_state["t1_conditions"] = [c for i, c in enumerate(conds) if i not in del_idx]
                    st.rerun()
            if bc2.button("전체 삭제", key="t1_delall", use_container_width=True):
                st.session_state["t1_conditions"] = []
                st.rerun()

        with st.container(border=True):
            st.markdown("**Weibull 시험시간비 적용 (선택)**")
            wb_apply = st.checkbox("적용", key="t1_wbapply")
            wc1, wc2 = st.columns(2)
            Rv = wc1.number_input("목표신뢰도 R", value=0.99, step=0.01, format="%.3f", key="t1_R")
            CL = wc2.number_input("신뢰수준 CL", value=0.5, step=0.05, format="%.3f", key="t1_CL")
            wc3, wc4 = st.columns(2)
            ns = wc3.number_input("샘플수 n", value=6.0, step=1.0, key="t1_n")
            beta_i = wc4.number_input("형상모수 β", value=2.0, step=0.1, key="t1_beta")

        ac1, ac2 = st.columns(2)
        run1 = ac1.button("계산 실행 ▶", type="primary", use_container_width=True, key="t1_run")

        def _t1_build_wb():
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "온도가속(Arrhenius)"
            bold = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="DCE6F1")
            d = st.session_state["t1_result"]
            ws.append(["활성화에너지 E(eV)", d["E"], "대표(목표)온도 Tref(℃)", d["Tref"]])
            ws.append([])
            ws.append(["조건온도(℃)", "조건시간(h)", "가속계수 AF", "등가시간(h)"])
            for c in ws[3]:
                c.font = bold; c.fill = header_fill
            for ta_, t_, af_, eq_ in d["rows"]:
                ws.append([ta_, t_, af_, eq_])
            ws.append([])
            ws.append(["총 등가시험시간(h)", d["total"]])
            if d["ratio"] is not None:
                ws.append(["Weibull 시험시간비", d["ratio"]])
                ws.append(["최종 가속시험시간(h)", d["final"]])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
            return wb

        with ac2:
            excel_download_button("엑셀로 내보내기 📊", "t1_xl", _t1_build_wb, "온도가속_계산결과.xlsx")

        if run1:
            if not conds:
                st.warning("조건을 1개 이상 추가해주세요.")
            else:
                rows_out = []
                total = 0.0
                for ta_, t_ in conds:
                    af_ = R.arrhenius_af(Tref, ta_, E)
                    eq_ = t_ * af_
                    total += eq_
                    rows_out.append((ta_, t_, af_, eq_))
                final_time = total
                ratio = None
                if wb_apply:
                    ratio = R.weibull_test_time_ratio(Rv, CL, ns, beta_i)
                    final_time = total * ratio
                st.session_state["t1_result"] = dict(E=E, Tref=Tref, rows=rows_out, total=total,
                                                      ratio=ratio, final=final_time)
                st.session_state["t1_xl_ready"] = True
                st.session_state["t1_xl_bytes"] = None

    with right:
        st.markdown("#### 결과")
        d = st.session_state.get("t1_result")
        if d is None:
            st.markdown("<div class='result-box'>조건을 입력하고 [계산 실행]을 눌러주세요.</div>", unsafe_allow_html=True)
        else:
            msg = f"총 등가시험시간(@{d['Tref']:.1f}℃) = {_fmt_num(d['total'])} h  ({_fmt_num(d['total']/24)} 일)"
            if d["ratio"] is not None:
                msg += f"\nWeibull 시험시간비 = {d['ratio']:.4f}"
                msg += f"\n최종 가속시험시간 = {_fmt_num(d['final'])} h  ({_fmt_num(d['final']/24)} 일)"
            st.markdown(f"<div class='result-box'>{msg}</div>", unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**계산 공식 (Arrhenius Model)**")
            st.markdown("""<div class='formula-box'>AF = exp[ (E / k) x (1/Tu - 1/Ta) ]

여기서  AF : 가속계수
        E  : 활성화에너지 (eV)
        k  : Boltzmann 상수 (8.6173e-05 eV/K)
        Tu : 사용(필드) 온도 (K)
        Ta : 가속(시험) 온도 (K)

등가시험시간 = 조건시간 / AF  (조건온도 -> Tref 로 환산)
최종 가속시험시간 = 등가시험시간 합계 x Weibull 시험시간비</div>""", unsafe_allow_html=True)

        if d is not None:
            detail = ["[조건별 상세]"]
            for ta_, t_, af_, eq_ in d["rows"]:
                detail.append(f"  {ta_:.1f}℃ / {t_:.1f}h  ->  AF={af_:,.3f}   등가시간={eq_:,.2f}h")
            st.markdown(f"<div class='detail-box'>{chr(10).join(detail)}</div>", unsafe_allow_html=True)


# ====================================================================
# 탭② 온습도가속 (Arrhenius-Peck Model)
# ====================================================================
with tab2:
    left, right = st.columns([2, 3])
    with left:
        st.markdown("#### ② 온습도가속 (Arrhenius-Peck Model)")

        with st.container(border=True):
            st.markdown("**입력값**")
            c1, c2, c3 = st.columns([2, 1, 2])
            with c1:
                E2 = st.number_input("활성화에너지 E (eV)", value=0.79, step=0.01, format="%.3f", key="t2_E")
            with c2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                with st.popover("DB"):
                    q2 = st.text_input("부품명 검색", key="t2_dbq")
                    rows2 = [r for r in R.EA_DB if q2.lower() in r[0].lower()] if q2 else R.EA_DB
                    st.dataframe(pd.DataFrame(rows2, columns=["부품", "활성화에너지(eV)", "출처", "비고"]),
                                 height=280, use_container_width=True)
                    names2 = [r[0] for r in rows2]
                    sel2 = st.selectbox("적용할 항목", ["-"] + names2, key="t2_dbsel")
                    if st.button("선택값 적용", key="t2_dbapply"):
                        if sel2 != "-":
                            row = next(r for r in rows2 if r[0] == sel2)
                            ev_text = str(row[1]).split("~")[0].split("-")[0].replace("eV", "").strip()
                            try:
                                st.session_state["t2_E"] = float(ev_text)
                                st.rerun()
                            except ValueError:
                                st.warning("이 항목은 단일 숫자값이 아니라 자동입력이 어렵습니다. 직접 입력해주세요.")
            with c3:
                n2 = st.number_input("습도항 지수 n", value=2.66, step=0.01, format="%.3f", key="t2_n")
            cc1, cc2 = st.columns(2)
            Tu2 = cc1.number_input("사용(필드) 온도 Tu(℃)", value=23.0, step=1.0, key="t2_Tu")
            Hu2 = cc2.number_input("사용(필드) 습도 Hu(%RH)", value=65.0, step=1.0, key="t2_Hu")

        with st.container(border=True):
            st.markdown("**가속시험 조건 (온도 / 습도 / 시간) 추가**")
            g1, g2, g3, g4 = st.columns([1, 1, 1, 1])
            ta2 = g1.number_input("가속온도(℃)", value=85.0, step=1.0, key="t2_ta")
            ha2 = g2.number_input("가속습도(%RH)", value=85.0, step=1.0, key="t2_ha")
            tt2 = g3.number_input("시간(h)", value=303.0, step=1.0, key="t2_tt")
            with g4:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("추가 +", key="t2_add"):
                    st.session_state.setdefault("t2_conditions", []).append((ta2, ha2, tt2))

        with st.container(border=True):
            st.markdown("**조건 목록**")
            conds2 = st.session_state.get("t2_conditions", [])
            del_idx2 = []
            if conds2:
                df2 = pd.DataFrame(conds2, columns=["가속온도(℃)", "가속습도(%RH)", "시간(h)"])
                st.dataframe(df2, use_container_width=True, height=180)
                del_idx2 = st.multiselect("삭제할 행 선택", list(range(len(conds2))), key="t2_delsel",
                                           format_func=lambda i: f"{i}: {conds2[i][0]}℃/{conds2[i][1]}%RH/{conds2[i][2]}h")
            else:
                st.caption("추가된 조건이 없습니다.")
            b1, b2 = st.columns(2)
            if b1.button("선택 삭제 🗑", key="t2_delbtn", use_container_width=True):
                if conds2 and del_idx2:
                    st.session_state["t2_conditions"] = [c for i, c in enumerate(conds2) if i not in del_idx2]
                    st.rerun()
            if b2.button("전체 삭제", key="t2_delall", use_container_width=True):
                st.session_state["t2_conditions"] = []
                st.rerun()

        with st.container(border=True):
            st.markdown("**Weibull 시험시간비 적용 (선택)**")
            wb_apply2 = st.checkbox("적용", key="t2_wbapply")
            w1, w2 = st.columns(2)
            Rv2 = w1.number_input("목표신뢰도 R", value=0.99, step=0.01, format="%.3f", key="t2_R")
            CL2 = w2.number_input("신뢰수준 CL", value=0.5, step=0.05, format="%.3f", key="t2_CL")
            w3, w4 = st.columns(2)
            ns2 = w3.number_input("샘플수 n", value=6.0, step=1.0, key="t2_ns")
            beta2 = w4.number_input("형상모수 β", value=2.0, step=0.1, key="t2_beta")

        a1, a2 = st.columns(2)
        run2 = a1.button("계산 실행 ▶", type="primary", use_container_width=True, key="t2_run")

        def _t2_build_wb():
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "온습도가속(Peck)"
            bold = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="DCE6F1")
            d = st.session_state["t2_result"]
            ws.append(["E(eV)", d["E"], "n", d["n"], "Tu(℃)", d["Tu"], "Hu(%RH)", d["Hu"]])
            ws.append([])
            ws.append(["가속온도(℃)", "가속습도(%RH)", "시간(h)", "AF", "등가시간(h)"])
            for c in ws[3]:
                c.font = bold; c.fill = header_fill
            for ta_, ha_, t_, af_, eq_ in d["rows"]:
                ws.append([ta_, ha_, t_, af_, eq_])
            ws.append([])
            ws.append(["총 등가시험시간(h)", d["total"]])
            if d["ratio"] is not None:
                ws.append(["Weibull 시험시간비", d["ratio"]])
                ws.append(["최종 가속시험시간(h)", d["final"]])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
            return wb

        with a2:
            excel_download_button("엑셀로 내보내기 📊", "t2_xl", _t2_build_wb, "온습도가속_계산결과.xlsx")

        if run2:
            if not conds2:
                st.warning("조건을 1개 이상 추가해주세요.")
            else:
                rows_out2 = []
                total2 = 0.0
                for ta_, ha_, t_ in conds2:
                    af_ = R.peck_af(Tu2, Hu2, ta_, ha_, E2, n2)
                    eq_ = t_ * af_
                    total2 += eq_
                    rows_out2.append((ta_, ha_, t_, af_, eq_))
                final2 = total2
                ratio2 = None
                if wb_apply2:
                    ratio2 = R.weibull_test_time_ratio(Rv2, CL2, ns2, beta2)
                    final2 = total2 * ratio2
                st.session_state["t2_result"] = dict(E=E2, n=n2, Tu=Tu2, Hu=Hu2, rows=rows_out2,
                                                      total=total2, ratio=ratio2, final=final2)
                st.session_state["t2_xl_ready"] = True
                st.session_state["t2_xl_bytes"] = None

    with right:
        st.markdown("#### 결과")
        d2 = st.session_state.get("t2_result")
        if d2 is None:
            st.markdown("<div class='result-box'>조건을 입력하고 [계산 실행]을 눌러주세요.</div>", unsafe_allow_html=True)
        else:
            msg2 = f"총 등가시험시간(@{d2['Tu']:.1f}℃/{d2['Hu']:.1f}%RH 기준) = {_fmt_num(d2['total'])} h  ({_fmt_num(d2['total']/24)} 일)"
            if d2["ratio"] is not None:
                msg2 += f"\nWeibull 시험시간비 = {d2['ratio']:.4f}"
                msg2 += f"\n최종 가속시험시간 = {_fmt_num(d2['final'])} h  ({_fmt_num(d2['final']/24)} 일)"
            st.markdown(f"<div class='result-box'>{msg2}</div>", unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**계산 공식 (Arrhenius-Peck Model)**")
            st.markdown("""<div class='formula-box'>AF = Lu/La = exp[(E/k) x (1/Tu - 1/Ta)] x (Hu/Ha)^(-n)

여기서  AF : 가속계수(Acceleration Factor)
        L  : 제품의 수명(h)
        E  : 활성화 에너지(Activation Energy), eV
        k  : Boltzmann 상수(=8.6173e-05 eV/K)
        T  : 절대 온도(K)
        H  : 상대 습도(% RH)
        n  : 습도항 지수
        첨자 a : 가속조건 / 첨자 u : 비가속조건 또는 사용자 환경조건

등가시험시간 = 조건시간 / AF
최종 가속시험시간 = 등가시험시간 합계 x Weibull 시험시간비</div>""", unsafe_allow_html=True)

        if d2 is not None:
            detail2 = ["[조건별 상세]"]
            for ta_, ha_, t_, af_, eq_ in d2["rows"]:
                detail2.append(f"  {ta_:.1f}℃/{ha_:.1f}%RH / {t_:.1f}h  ->  AF={af_:,.2f}   등가시간={eq_:,.2f}h")
            st.markdown(f"<div class='detail-box'>{chr(10).join(detail2)}</div>", unsafe_allow_html=True)


# ====================================================================
# 탭③ 열피로가속 (Thermal Cycling)
# ====================================================================
with tab3:
    left, right = st.columns([1, 1])
    with left:
        st.markdown("#### ③ 열피로가속 (Thermal Cycling)")
        st.markdown("<span class='note-red'>※ 챔버 승온형 온도사이클(Thermal Cycling) 전용입니다. "
                    "(엘리베이터형 Thermal Shock 시험은 전환시간이 거의 0으로, 별도 모델이 필요합니다)</span>",
                    unsafe_allow_html=True)

        pc1, pc2 = st.columns(2)
        with pc1:
            with st.container(border=True):
                st.markdown("**사용조건(필드) 사이클 프로파일**")
                f_low = st.number_input("저온(℃)", value=-40.0, key="t3_f_low")
                f_low_dwell = st.number_input("저온유지시간(분)", value=10.0, key="t3_f_lowd")
                f_ramp_up = st.number_input("승온시간(분, 저온→고온)", value=30.0, key="t3_f_rampu")
                f_high = st.number_input("고온(℃)", value=85.0, key="t3_f_high")
                f_high_dwell = st.number_input("고온유지시간(분)", value=10.0, key="t3_f_highd")
                f_ramp_down = st.number_input("하강시간(분, 고온→저온)", value=30.0, key="t3_f_rampd")
                f_target_cycle = st.number_input("필드 목표 사이클수", value=1000.0, key="t3_f_cycle")
        with pc2:
            with st.container(border=True):
                st.markdown("**시험조건 사이클 프로파일**")
                s_low = st.number_input("저온(℃)", value=-40.0, key="t3_s_low")
                s_low_dwell = st.number_input("저온유지시간(분)", value=10.0, key="t3_s_lowd")
                s_ramp_up = st.number_input("승온시간(분, 저온→고온)", value=30.0, key="t3_s_rampu")
                s_high = st.number_input("고온(℃)", value=125.0, key="t3_s_high")
                s_high_dwell = st.number_input("고온유지시간(분)", value=40.0, key="t3_s_highd")
                s_ramp_down = st.number_input("하강시간(분, 고온→저온)", value=10.0, key="t3_s_rampd")

        st.markdown("<span class='note-gray'>예) -40℃(30분) ↔ 125℃(30분), 승온/하강 각 10분  형태로 입력하시면 됩니다.</span>",
                    unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**가속 모델 선택**")
            gc1, gc2 = st.columns([1, 1])
            with gc1:
                model = st.radio("모델", ["Coffin-Manson (단순, ΔT만 반영)",
                                         "Modified Norris-Landzberg (정밀, dwell/ramp/온도 반영)"],
                                  key="t3_model", label_visibility="collapsed")
                m_col1, m_col2 = st.columns([1, 1])
                m_val = m_col1.number_input("m지수", value=2.5, step=0.1, key="t3_m")
                with m_col2:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    with st.popover("m지수 가이드"):
                        st.dataframe(pd.DataFrame(R.M_GUIDE_DB, columns=["재질/메커니즘", "m지수", "대상 부품(예)"]),
                                     height=250, use_container_width=True)
                        labels = [row[0] for row in R.M_GUIDE_DB]
                        sel_m = st.selectbox("적용할 항목", ["-"] + labels, key="t3_mguide_sel")
                        if st.button("선택값 적용", key="t3_mguide_apply"):
                            if sel_m != "-":
                                row = next(r for r in R.M_GUIDE_DB if r[0] == sel_m)
                                st.session_state["t3_m"] = float(row[1])
                                st.rerun()
            with gc2:
                st.markdown("**언제 어떤 모델을 써야 하나요?**")
                st.markdown(f"<div class='formula-box' style='background-color:#fbf7ec; height:230px; overflow-y:auto;'>{R.MODEL_GUIDE_TEXT}</div>",
                            unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**Weibull 시험시간비 적용 (선택)**")
            wb_apply3 = st.checkbox("적용", key="t3_wbapply")
            w1, w2 = st.columns(2)
            Rv3 = w1.number_input("목표신뢰도 R", value=0.99, step=0.01, format="%.3f", key="t3_R")
            CL3 = w2.number_input("신뢰수준 CL", value=0.5, step=0.05, format="%.3f", key="t3_CL")
            w3, w4 = st.columns(2)
            ns3 = w3.number_input("샘플수 n", value=6.0, step=1.0, key="t3_ns")
            beta3 = w4.number_input("형상모수 β", value=2.0, step=0.1, key="t3_beta")

        a1, a2 = st.columns(2)
        run3 = a1.button("계산 실행 ▶", type="primary", use_container_width=True, key="t3_run")

        def _t3_build_wb():
            import openpyxl
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "열피로가속(Thermal Cycling)"
            bold = Font(bold=True)
            d = st.session_state["t3_result"]
            ws.append(["모델", d["model_name"]])
            ws.append(["ΔT_field(℃)", d["dT_field"], "ΔT_test(℃)", d["dT_test"]])
            ws.append(["가속계수 AF", d["af"]])
            ws.append(["필요 시험 사이클수", d["need_cycle"]])
            ws.append(["1cycle 시험시간(분)", d["cycle_time_test"]])
            ws.append(["총 소요시간(시간)", d["total_hour"], "총 소요시간(일)", d["total_day"]])
            if d["ratio"] is not None:
                ws.append(["Weibull 시험시간비", d["ratio"]])
                ws.append(["최종 가속시험 사이클수", d["final_cycle"]])
                ws.append(["최종 가속시험시간(시간)", d["final_hour"]])
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.column == 1:
                        cell.font = bold
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 24
            return wb

        with a2:
            excel_download_button("엑셀로 내보내기 📊", "t3_xl", _t3_build_wb, "열피로가속_계산결과.xlsx")

        if run3:
            dT_field = abs(f_high - f_low)
            dT_test = abs(s_high - s_low)
            cycle_time_field = R.profile_cycle_time_min(f_low_dwell, f_ramp_up, f_high_dwell, f_ramp_down)
            cycle_time_test = R.profile_cycle_time_min(s_low_dwell, s_ramp_up, s_high_dwell, s_ramp_down)

            is_coffin = model.startswith("Coffin")
            if is_coffin:
                af3 = R.coffin_manson_af(dT_field, dT_test, m_val)
                model_name = "Coffin-Manson"
            else:
                ramp_test_rate = dT_test / s_ramp_up if s_ramp_up > 0 else 0.0001
                af3 = R.norris_landzberg_af(dT_field, dT_test, f_high_dwell, s_high_dwell,
                                            ramp_test_rate, f_high, s_high)
                model_name = "Modified Norris-Landzberg"

            need_cycle = f_target_cycle / af3 if af3 > 0 else float("inf")
            total_min = need_cycle * cycle_time_test
            total_hour = total_min / 60
            total_day = total_hour / 24

            final_cycle = need_cycle
            final_min = total_min
            ratio3 = None
            if wb_apply3:
                ratio3 = R.weibull_test_time_ratio(Rv3, CL3, ns3, beta3)
                final_cycle = need_cycle * ratio3
                final_min = final_cycle * cycle_time_test

            st.session_state["t3_result"] = dict(
                dT_field=dT_field, dT_test=dT_test, cycle_time_field=cycle_time_field,
                cycle_time_test=cycle_time_test, af=af3, need_cycle=need_cycle,
                total_min=total_min, total_hour=total_hour, total_day=total_day,
                ratio=ratio3, final_cycle=final_cycle, final_hour=final_min / 60,
                model_name=model_name)
            st.session_state["t3_xl_ready"] = True
            st.session_state["t3_xl_bytes"] = None

    with right:
        st.markdown("#### 결과")
        d3 = st.session_state.get("t3_result")
        if d3 is None:
            st.markdown("<div class='result-box'>프로파일을 입력하고 [계산 실행]을 눌러주세요.</div>", unsafe_allow_html=True)
        else:
            lines3 = [
                f"ΔT(필드/시험) = {d3['dT_field']:.1f}℃ / {d3['dT_test']:.1f}℃",
                f"1cycle 시간(필드/시험) = {d3['cycle_time_field']:.1f}분 / {d3['cycle_time_test']:.1f}분",
                "",
                f"가속계수 AF = {d3['af']:,.3f}",
                f"필요 시험 사이클수 = {d3['need_cycle']:,.2f} cycle",
                f"총 소요시간 = {d3['total_min']:,.1f}분 = {d3['total_hour']:,.2f}시간 = {d3['total_day']:,.2f}일",
            ]
            if d3["ratio"] is not None:
                lines3.append("")
                lines3.append(f"Weibull 시험시간비 = {d3['ratio']:.4f}")
                lines3.append(f"최종 가속시험 사이클수 = {d3['final_cycle']:,.2f} cycle "
                               f"({d3['final_hour']:,.2f}시간 = {d3['final_hour']/24:,.2f}일)")
            st.markdown(f"<div class='result-box'>{chr(10).join(lines3)}</div>", unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**계산 공식**")
            if st.session_state.get("t3_result") and st.session_state["t3_result"]["model_name"] == "Coffin-Manson" \
                    or not st.session_state.get("t3_result"):
                formula3 = ("[Coffin-Manson Model]\n\n"
                            "AF = ( ΔT_test / ΔT_field ) ^ m\n\n"
                            "여기서  AF : 가속계수\n"
                            "        ΔT : 온도변화폭(사이클 최고온도-최저온도)\n"
                            "        m  : 재료/구조에 따른 경험적 지수 (보통 1.9~5)\n\n"
                            "필요 시험 사이클수 = 필드 목표 사이클수 / AF\n"
                            "1cycle 시간 = 저온유지 + 승온시간 + 고온유지 + 하강시간\n"
                            "총 소요시간 = 필요 시험 사이클수 x 1cycle 시간")
            else:
                formula3 = ("[Modified Norris-Landzberg Model]\n\n"
                            "AF = (ΔT_test/ΔT_field)^2.65\n"
                            "     x (Dwell_test/Dwell_field)^0.136\n"
                            "     x (1.22 x RampRate_test^-0.0757)\n"
                            "     x exp[2185 x (1/Tmax_field(K) - 1/Tmax_test(K))]\n\n"
                            "여기서  Dwell : 고온유지시간,  RampRate : 승온속도(℃/분)\n"
                            "        Tmax  : 사이클 최고온도(K)\n"
                            "        (계수 n=2.65, m=0.136, Ea/k=2185 는 Pb-free 솔더 접합 문헌 기준 근사값)\n\n"
                            "필요 시험 사이클수 = 필드 목표 사이클수 / AF\n"
                            "1cycle 시간 = 저온유지 + 승온시간 + 고온유지 + 하강시간\n"
                            "총 소요시간 = 필요 시험 사이클수 x 1cycle 시간")
            st.markdown(f"<div class='formula-box'>{formula3}</div>", unsafe_allow_html=True)


# ====================================================================
# 탭④ Weibull 시험시간비 계산기
# ====================================================================
with tab4:
    left, right = st.columns([2, 3])
    with left:
        st.markdown("#### ④ Weibull 시험시간비 계산기")

        with st.container(border=True):
            st.markdown("**입력값**")
            Rv4 = st.number_input("목표신뢰도 R", value=0.99, step=0.01, format="%.3f", key="t4_R")
            CL4 = st.number_input("신뢰수준 CL", value=0.5, step=0.05, format="%.3f", key="t4_CL")
            ns4 = st.number_input("샘플수 n", value=6.0, step=1.0, key="t4_n")
            beta4 = st.number_input("형상모수 β", value=2.0, step=0.1, key="t4_beta")

        with st.popover("Beta 참고DB 열기", use_container_width=True):
            q4 = st.text_input("부품명 검색", key="t4_dbq")
            wrows = [r for r in R.WEIBULL_DB if q4.lower() in r[1].lower()] if q4 else R.WEIBULL_DB
            st.dataframe(pd.DataFrame(wrows, columns=["분류", "항목", "Beta Low", "Beta Typ", "Beta High",
                                                       "Eta Low", "Eta Typ", "Eta High"]),
                         height=350, use_container_width=True)
            st.caption("※ 참고용 DB입니다 (원본과 동일하게 값 자동적용 기능은 제공하지 않습니다).")

        run4 = st.button("계산 실행 ▶", type="primary", use_container_width=True, key="t4_run")
        if run4:
            try:
                ratio4 = R.weibull_test_time_ratio(Rv4, CL4, ns4, beta4)
                st.session_state["t4_result"] = ratio4
            except (ValueError, ZeroDivisionError):
                st.warning("값을 확인해주세요.")

    with right:
        st.markdown("#### 결과")
        r4 = st.session_state.get("t4_result")
        if r4 is None:
            st.markdown("<div class='result-box'>값을 입력하고 [계산 실행]을 눌러주세요.</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='result-box'>시험시간비 = {r4:.4f}</div>", unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**계산 공식**")
            st.markdown("""<div class='formula-box'>시험시간비 = [ (-ln CL) / (n x -ln R) ] ^ (1/β)

최종 가속시험시간(또는 사이클수) = 등가시험시간(사이클수) x 시험시간비</div>""", unsafe_allow_html=True)


# ====================================================================
# 탭⑤ 수명데이터 분석 (Weibull MLE)
# ====================================================================
with tab5:
    left, right = st.columns([2, 3])
    with left:
        st.markdown("#### ⑤ 수명데이터 분석 (Weibull)")

        with st.container(border=True):
            st.markdown("**고장/미고장 데이터 입력**")
            r1c1, r1c2, r1c3 = st.columns([1, 1, 1])
            t5_time = r1c1.number_input("시간(h 또는 cycle)", value=0.0, step=1.0, key="t5_time")
            t5_status = r1c2.radio("상태", ["고장", "미고장"], key="t5_status", horizontal=True)
            with r1c3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("추가 +", key="t5_add"):
                    if t5_time > 0:
                        st.session_state.setdefault("t5_rows", []).append((t5_time, t5_status == "고장"))
                    else:
                        st.warning("시간을 올바르게 입력해주세요. (0보다 큰 값)")
            st.caption("※ 미고장 = 관찰(시험) 종료 시점까지 고장이 발생하지 않은 데이터")

            rows5 = st.session_state.get("t5_rows", [])
            del_idx5 = []
            if rows5:
                df5 = pd.DataFrame([(t, "고장" if f else "미고장") for t, f in rows5], columns=["시간", "상태"])
                st.dataframe(df5, use_container_width=True, height=220)
                del_idx5 = st.multiselect("삭제할 행 선택", list(range(len(rows5))), key="t5_delsel",
                                           format_func=lambda i: f"{i}: {rows5[i][0]} / {'고장' if rows5[i][1] else '미고장'}")
            else:
                st.caption("입력된 데이터가 없습니다.")

            bc1, bc2, bc3 = st.columns(3)
            if bc1.button("선택 삭제", key="t5_delbtn", use_container_width=True):
                if rows5 and del_idx5:
                    st.session_state["t5_rows"] = [r for i, r in enumerate(rows5) if i not in del_idx5]
                    st.rerun()
            if bc2.button("전체 삭제", key="t5_delall", use_container_width=True):
                st.session_state["t5_rows"] = []
                st.session_state["t5_fit"] = None
                st.rerun()
            if bc3.button("예시데이터 불러오기", key="t5_example", use_container_width=True):
                example = [(185, True), (260, True), (312, True), (355, True), (398, True),
                           (430, True), (470, True), (520, True), (610, True),
                           (700, False), (700, False), (700, False), (700, False), (700, False), (700, False)]
                st.session_state["t5_rows"] = example
                st.rerun()

        run5 = st.button("Weibull 적합(MLE) 실행 ▶", type="primary", use_container_width=True, key="t5_fitbtn")

        def _t5_build_wb():
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "수명데이터"
            header_fill = PatternFill("solid", fgColor="DDEBF7")
            ws.append(["시간", "상태"])
            for c in ws[1]:
                c.font = Font(bold=True); c.fill = header_fill
            for t, f in st.session_state.get("t5_rows", []):
                ws.append([t, "고장" if f else "미고장"])
            ws2 = wb.create_sheet("적합결과")
            ws2.append(["항목", "값"])
            for c in ws2[1]:
                c.font = Font(bold=True); c.fill = header_fill
            fit = st.session_state.get("t5_fit")
            if fit is not None:
                beta_, eta_ = fit
                ws2.append(["형상모수 β", beta_])
                ws2.append(["척도모수 η", eta_])
                ws2.append(["MTTF(평균수명)", R.weibull_mttf(beta_, eta_)])
                ws2.append(["B10 수명", R.weibull_bpercentile(beta_, eta_, 10)])
                ws2.append(["B1 수명", R.weibull_bpercentile(beta_, eta_, 1)])
            else:
                ws2.append(["안내", "적합을 먼저 실행해주세요."])
            return wb

        excel_download_button("엑셀로 내보내기", "t5_xl", _t5_build_wb, "수명데이터분석_결과.xlsx")

        if run5:
            rows5 = st.session_state.get("t5_rows", [])
            if len(rows5) < 3:
                st.warning("데이터를 3개 이상 입력해주세요 (고장 데이터 2개 이상 포함).")
            else:
                times = [r[0] for r in rows5]
                is_failure = [r[1] for r in rows5]
                if sum(is_failure) < 2:
                    st.warning("고장(Failure) 데이터가 2개 이상 필요합니다.")
                else:
                    result = R.fit_weibull_mle(times, is_failure)
                    if result is None:
                        st.warning("Weibull 적합에 실패했습니다. 데이터를 확인해주세요.")
                    else:
                        st.session_state["t5_fit"] = result
                        st.session_state["t5_xl_ready"] = True
                        st.session_state["t5_xl_bytes"] = None
                        st.rerun()

    with right:
        st.markdown("#### 적합 결과")
        fit = st.session_state.get("t5_fit")
        rows5 = st.session_state.get("t5_rows", [])
        if fit is None:
            st.markdown("<div class='result-box'>데이터를 입력하고 [Weibull 적합(MLE) 실행]을 눌러주세요. "
                        "(고장 데이터 2개 이상 필요)</div>", unsafe_allow_html=True)
        else:
            beta_, eta_ = fit
            times = [r[0] for r in rows5]
            is_failure = [r[1] for r in rows5]
            n_fail = sum(is_failure)
            n_total = len(rows5)
            if beta_ < 1.0:
                interp = "β<1 : 초기고장(Infant Mortality) 특성 - 시간이 지날수록 고장률이 감소"
            elif beta_ < 1.3:
                interp = "β≈1 : 우발고장(Random Failure) 특성 - 고장률이 시간과 거의 무관"
            else:
                interp = "β>1 : 마모성고장(Wear-out) 특성 - 시간이 지날수록 고장률이 증가"
            msg5 = (f"형상모수 β = {beta_:.4f}   척도모수 η = {eta_:,.2f}\n"
                    f"(총 {n_total}개 데이터 : 고장 {n_fail}개 / 미고장 {n_total - n_fail}개)\n"
                    f"{interp}")
            st.markdown(f"<div class='result-box'>{msg5}</div>", unsafe_allow_html=True)

        if st.button("이 β값을 ④ Weibull 탭으로 보내기 →", key="t5_send", disabled=(fit is None)):
            beta_, eta_ = fit
            st.session_state["t4_beta"] = float(f"{beta_:.4f}")
            st.success(f"β = {beta_:.4f} 값을 ④ Weibull 시험시간비 탭으로 보냈습니다. "
                       "상단의 '④ Weibull 시험시간비' 탭을 클릭해서 확인해주세요.")

        with st.container(border=True):
            hc1, hc2 = st.columns([4, 1])
            hc1.markdown("**대표수명 지표 (MTTF / B10 / B1)**")
            with hc2:
                with st.popover("ⓘ 왜 필요한가?"):
                    st.markdown(
                        "β(형상모수)와 η(척도모수)만으로는 '이 제품이 실제로 몇 시간/사이클을 버티는지'가 "
                        "바로 감이 오지 않기 때문에, 실무에서 바로 쓸 수 있는 대표수명값으로 변환한 지표입니다.\n\n"
                        "· MTTF (평균수명)\n"
                        "  전체 제품의 평균적인 고장 시점입니다. '평균적으로 이 정도 시간에 고장난다'는 의미이며,\n"
                        "  설계수명과 비교해 여유(마진)가 있는지 확인하는 데 씁니다.\n\n"
                        "· B10 수명\n"
                        "  전체 중 10%가 고장 나는 시점입니다. 자동차/전자부품 업계에서 널리 쓰이는\n"
                        "  표준 신뢰성 판정 지표로, '고객에게 보증할 수명'을 정할 때 기준으로 많이 사용합니다.\n\n"
                        "· B1 수명\n"
                        "  전체 중 1%가 고장 나는 시점입니다. B10보다 더 보수적인(엄격한) 기준이며,\n"
                        "  안전과 직결되는 부품이나 높은 신뢰성이 요구되는 부품에 사용합니다.\n\n"
                        "즉 세 지표 모두 β/η 모델을 '몇 % 고장 시점이 언제인지'로 쉽게 풀어서\n"
                        "리포트나 고객 스펙 비교에 바로 활용할 수 있게 해주는 값입니다.")
            if fit is not None:
                beta_, eta_ = fit
                mttf_ = R.weibull_mttf(beta_, eta_)
                b10_ = R.weibull_bpercentile(beta_, eta_, 10)
                b1_ = R.weibull_bpercentile(beta_, eta_, 1)
                st.markdown(f"<div class='detail-box'>MTTF(평균수명)   = {mttf_:,.2f}\n"
                            f"B10 수명(F=10%) = {b10_:,.2f}\nB1 수명(F=1%)   = {b1_:,.2f}</div>",
                            unsafe_allow_html=True)
            else:
                st.markdown("<div class='detail-box'>-</div>", unsafe_allow_html=True)

        with st.container(border=True):
            hc1, hc2 = st.columns([4, 1])
            hc1.markdown("**임의 시점 신뢰도 계산**")
            with hc2:
                with st.popover("ⓘ 왜 필요한가?"):
                    st.markdown(
                        "이 계산은 '시험을 얼마나 오래 했는지'와는 무관하게, 사용자가 알고 싶은\n"
                        "특정 시점(t) 하나를 골라서 그 시점의 신뢰도를 계산해주는 기능입니다.\n\n"
                        "예를 들어 시험 데이터를 1500시간까지 수집해 β/η을 추정해 놓은 상태에서:\n"
                        "  · t=200을 입력하면 → '200시간 시점까지 생존할 확률'\n"
                        "  · t=1500을 입력하면 → '시험 종료 시점까지 생존할 확률'\n"
                        "  · t=87600(필드수명 10년)을 입력하면 → '실제 필드수명 시점에서의 예상 생존율'\n"
                        "을 즉시 확인할 수 있습니다.\n\n"
                        "실무에서는 초기불량 구간(짧은 t) 확인, 시험 통과 확률의 사후 검증(시험시간과 같은 t),\n"
                        "그리고 필드 목표수명 시점의 신뢰도 예측(긴 t) 등의 용도로 가장 많이 사용합니다.\n\n"
                        "※ 주의: t가 실제 관측한 데이터 범위를 크게 벗어나면(외삽, extrapolation)\n"
                        "모델의 예측 오차가 커질 수 있으므로, 참고값으로만 활용하시기 바랍니다.")
            rc1, rc2 = st.columns([2, 1])
            t5_t = rc1.number_input("임의 시점 t =", value=1000.0, step=10.0, key="t5_t")
            with rc2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                calc_rt = st.button("계산", key="t5_calcrt", disabled=(fit is None))
            if calc_rt and fit is not None:
                beta_, eta_ = fit
                r_ = R.weibull_reliability(t5_t, beta_, eta_)
                f_ = 1 - r_
                h_ = R.weibull_failure_rate(t5_t, beta_, eta_)
                st.session_state["t5_rt_text"] = (
                    f"R({t5_t:g}) = {r_*100:.3f} %   (t시점까지 생존할 확률)\n"
                    f"F({t5_t:g}) = {f_*100:.3f} %   (t시점까지 고장날 확률)\n"
                    f"h({t5_t:g}) = {h_:.6f}   (t시점에서의 고장률)")
            st.markdown(f"<div class='detail-box'>{st.session_state.get('t5_rt_text', '-')}</div>",
                        unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("**Weibull 확률도표**")
            if fit is not None:
                beta_, eta_ = fit
                times = [r[0] for r in rows5]
                is_failure = [r[1] for r in rows5]
                ranks = R.weibull_rank_adjustment(times, is_failure)
                fig, ax = plt.subplots(figsize=(6, 4.6), dpi=100)
                if ranks:
                    xs = [math.log(t) for t, _ in ranks]
                    ys = [math.log(-math.log(1 - mr)) for _, mr in ranks]
                    ax.scatter(xs, ys, color="#2f6fed", label="관측 데이터(고장)")
                tmin, tmax = min(times), max(times)
                t_line = [tmin * (tmax / tmin) ** (i / 50.0) if tmin > 0 else tmax * i / 50.0 for i in range(51)]
                x_line = [math.log(t) for t in t_line if t > 0]
                y_line = [beta_ * (math.log(t) - math.log(eta_)) for t in t_line if t > 0]
                ax.plot(x_line, y_line, color="#e2483a", label=f"MLE 적합선 (β={beta_:.2f}, η={eta_:,.1f})")
                ax.set_xlabel("ln(시간)")
                ax.set_ylabel("ln(-ln(1-F))")
                ax.set_title("Weibull 확률도표")
                ax.legend(fontsize=8)
                ax.grid(alpha=0.3)
                fig.tight_layout()
                st.pyplot(fig)
            else:
                st.caption("적합을 실행하면 확률도표가 표시됩니다.")
